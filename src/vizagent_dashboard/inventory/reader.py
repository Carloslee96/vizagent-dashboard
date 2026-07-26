"""安全读取 CSV/XLSX，并生成确定性 DataInventory。"""

from __future__ import annotations

import csv
import hashlib
import math
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from vizagent_dashboard.inventory.spec import ColumnInfo, DataInventory, SheetInfo


@dataclass(frozen=True)
class InputPolicy:
    max_file_bytes: int = 50 * 1024 * 1024
    max_uncompressed_bytes: int = 250 * 1024 * 1024
    max_sheets: int = 50
    max_rows_per_sheet: int = 100_000
    max_columns: int = 500
    max_cell_chars: int = 10_000


DEFAULT_POLICY = InputPolicy()


def read_file(
    path: str | Path,
    policy: InputPolicy = DEFAULT_POLICY,
) -> dict[str, list[dict[str, Any]]]:
    """读取白名单数据文件，不执行公式、宏、外部链接或嵌入对象。"""

    source = Path(path).resolve()
    if not source.is_file():
        raise FileNotFoundError(f"文件不存在: {source}")
    if source.stat().st_size > policy.max_file_bytes:
        raise ValueError(f"文件超过 {policy.max_file_bytes // (1024 * 1024)}MB 限制")

    suffix = source.suffix.lower()
    if suffix == ".csv":
        return {"Sheet1": _read_csv(source, policy)}
    if suffix == ".xlsx":
        _check_xlsx_archive(source, policy)
        return _read_excel(source, policy)
    if suffix == ".xls":
        raise ValueError("暂不支持旧版 .xls，请另存为 .xlsx 后重试")
    raise ValueError(f"不支持的文件格式: {suffix}（仅支持 .csv / .xlsx）")


def inventory_file(
    path: str | Path,
    policy: InputPolicy = DEFAULT_POLICY,
) -> tuple[DataInventory, dict[str, list[dict[str, Any]]]]:
    """读取文件并返回 Inventory 与逐 Sheet 原始数据。"""

    source = Path(path).resolve()
    sheets = read_file(source, policy)
    sheet_infos: list[SheetInfo] = []
    for sheet_name, rows in sheets.items():
        column_names: list[str] = []
        for row in rows:
            for name in row:
                if name not in column_names:
                    column_names.append(name)
        columns = [_describe_column(name, rows) for name in column_names]
        sheet_infos.append(SheetInfo(name=sheet_name, row_count=len(rows), columns=columns))

    digest = hashlib.sha256()
    with source.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)

    inventory = DataInventory(
        source_path=str(source),
        source_sha256=digest.hexdigest(),
        total_rows=sum(len(rows) for rows in sheets.values()),
        sheets=sheet_infos,
    )
    return inventory, sheets


def _check_xlsx_archive(path: Path, policy: InputPolicy) -> None:
    if not zipfile.is_zipfile(path):
        raise ValueError("XLSX 文件结构无效")
    with zipfile.ZipFile(path) as archive:
        total = 0
        for info in archive.infolist():
            name = info.filename.replace("\\", "/")
            if name.startswith("/") or "../" in name.split("/"):
                raise ValueError("XLSX 包含不安全路径")
            total += info.file_size
            if total > policy.max_uncompressed_bytes:
                raise ValueError("XLSX 解压后体积超过安全限制")


def _read_csv(path: Path, policy: InputPolicy) -> list[dict[str, Any]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        try:
            raw_headers = next(reader)
        except StopIteration:
            return []
        headers = _normalize_headers(raw_headers, policy)
        rows: list[dict[str, Any]] = []
        for row_index, values in enumerate(reader, start=2):
            if len(rows) >= policy.max_rows_per_sheet:
                raise ValueError(f"CSV 超过 {policy.max_rows_per_sheet} 行限制")
            record = {
                header: _safe_cell(value, policy, row_index, header)
                for header, value in zip(headers, values)
                if header
            }
            if any(value not in (None, "") for value in record.values()):
                rows.append(record)
        return rows


def _read_excel(path: Path, policy: InputPolicy) -> dict[str, list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("读取 XLSX 需要安装 openpyxl") from exc

    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        if len(workbook.sheetnames) > policy.max_sheets:
            raise ValueError(f"工作表数量超过 {policy.max_sheets} 个限制")

        sheets: dict[str, list[dict[str, Any]]] = {}
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            iterator = worksheet.iter_rows(values_only=True)
            try:
                raw_headers = next(iterator)
            except StopIteration:
                sheets[sheet_name] = []
                continue
            headers = _normalize_headers(raw_headers, policy)
            rows: list[dict[str, Any]] = []
            for row_index, values in enumerate(iterator, start=2):
                if len(rows) >= policy.max_rows_per_sheet:
                    raise ValueError(f"工作表“{sheet_name}”超过 {policy.max_rows_per_sheet} 行限制")
                record = {
                    header: _safe_cell(value, policy, row_index, header)
                    for header, value in zip(headers, values)
                    if header
                }
                if any(value not in (None, "") for value in record.values()):
                    rows.append(record)
            sheets[sheet_name] = rows
        return sheets
    finally:
        workbook.close()


def _normalize_headers(raw_headers: Any, policy: InputPolicy) -> list[str]:
    values = list(raw_headers)[: policy.max_columns + 1]
    if len(values) > policy.max_columns:
        raise ValueError(f"列数超过 {policy.max_columns} 列限制")
    seen: dict[str, int] = {}
    headers: list[str] = []
    for index, raw in enumerate(values, start=1):
        base = str(raw).strip() if raw is not None else ""
        if not base:
            headers.append("")
            continue
        count = seen.get(base, 0) + 1
        seen[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    return headers


def _safe_cell(value: Any, policy: InputPolicy, row_index: int, header: str) -> Any:
    if isinstance(value, str) and len(value) > policy.max_cell_chars:
        raise ValueError(f"第 {row_index} 行字段“{header}”超过单元格长度限制")
    return value


def _describe_column(name: str, rows: list[dict[str, Any]]) -> ColumnInfo:
    values = [row.get(name) for row in rows]
    present = [value for value in values if value not in (None, "")]
    dtype = _infer_dtype(present)
    unique_serialized: list[Any] = []
    seen: set[str] = set()
    for value in present:
        key = repr(value)
        if key not in seen:
            seen.add(key)
            if len(unique_serialized) < 8:
                unique_serialized.append(value)

    minimum: float | str | None = None
    maximum: float | str | None = None
    if dtype == "numeric":
        numeric = [_to_number(value) for value in present]
        finite = [value for value in numeric if value is not None and math.isfinite(value)]
        if finite:
            minimum, maximum = min(finite), max(finite)
    elif dtype == "date" and present:
        normalized = [value.isoformat() if hasattr(value, "isoformat") else str(value) for value in present]
        minimum, maximum = min(normalized), max(normalized)

    return ColumnInfo(
        name=name,
        dtype=dtype,
        null_count=len(values) - len(present),
        unique_count=len(seen),
        sample_values=unique_serialized,
        minimum=minimum,
        maximum=maximum,
    )


def _infer_dtype(values: list[Any]) -> str:
    if not values:
        return "empty"
    if all(isinstance(value, (datetime, date)) for value in values):
        return "date"
    numeric = sum(_to_number(value) is not None for value in values)
    if numeric / len(values) >= 0.8:
        return "numeric"
    unique = len({repr(value) for value in values})
    if unique <= min(30, max(8, len(values) // 2)):
        return "categorical"
    return "text"


def _to_number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").replace("¥", "").replace("$", "").replace("%", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None
